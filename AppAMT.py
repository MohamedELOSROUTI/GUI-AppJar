from appJar import gui
from openpyxl import load_workbook, Workbook
import clx.xms
import requests
import re
import os
from datetime import datetime

def Convert(string):
# Fonction utilisée pour le formatage de charactères
    li = list(string.split("-")) 
    return li 

def findDiff(d1, d2, path=""):

# Fonction utilisée pour détecter les changements entre deux listes de bénéficiaires pour deux bénévoles différents.
# Inputs:
#     1) d1: Dictionnaire contenant coordonnées des bénéficiares accordées au bénévole x
#     2) d2: Dictionnaire contenant coordonnées des bénéficiaires accordées au bénévole y
# Outputs:
#     path : Liste dont le dernier élement est le bénéficiares manquants (présent dans x mais absent dans y)

    global benevole
    for k in d1:
        if (k not in d2):
            pass
        else:
            if type(d1[k]) is dict:
                if path == "":
                    path = k
                else:
                    path = path + "->" + k
                findDiff(d1[k],d2[k], path)
            else:
                if d1[k] != d2[k]:
                    benevole = path
def readAPI():
    # Fonction qui lit les identifiants API contenus dans Sinch.txt
    # Retourne l'id et le token
    try:
        with open("Sinch.txt", "r") as file:
            for line in file:
                if 'service_plan_id' in line:
                    l = len(line)
                    first_index = line.find("'")+1
                    last_index = line[first_index+1:].find("'")+first_index
                    service_plan_id = line[first_index:last_index]
                    print(service_plan_id)

                elif 'token' in line:
                    l = len(line)
                    first_index = line.find("'")+1
                    last_index = line[first_index+1:].find("'")+first_index
                    token = line[first_index:last_index]
                    print(token)
                else:
                    pass
            return service_plan_id, token
    except:
        app.infoBox("Erreur", "Le fichier Sinch (texte) contenant le service_plan_id et le token n'est pas présent !", parent=None)
        app.stop()

def press(btn):
# Fonction qui s'active lorsque le bouton envoie de message est pressé.
# Envoie un message au bénévole grâce à l'API inch.
# Un fichier texte contenant l'historique des messages envoyé est crée.

    global message
    global phone_numbers_benev_dic
    global now_
    global now
    # Si le dossier Historique des messages n'existe pas, pensez à le créer !
    if not os.path.exists('Historique des messages'):
        os.makedirs('Historique des messages')
    for index_ben in range(0,len(benevoles)):
        if btn == "Envoyer un message au bénévole "+str(index_ben+1):
            service_plan_id, token = readAPI()
            client = clx.xms.Client(service_plan_id=str(service_plan_id), token=str(token))

            create = clx.xms.api.MtBatchTextSmsCreate()
            create.sender = '447537404817'
            print(phone_numbers_benev_dic[benevoles[index_ben]])
            create.recipients = {str(phone_numbers_benev_dic[benevoles[index_ben]])}
            create.body = message[index_ben].replace("Message à envoyer à "+benevoles[index_ben]+":\n\n","")
            try:
                batch = client.create_batch(create)
                app.setMessage("Message bénévole "+str(index_ben+1),"Message envoyé à "+str(benevoles[index_ben]))
                file = open("Historique des messages/"+now_+".txt","a")
                file.write(message[index_ben].replace("Message à envoyer à "+benevoles[index_ben]+":\n\n",""))
                file.close()
            except (requests.exceptions.RequestException,
                clx.xms.exceptions.ApiException) as ex:

                app.setMessage("Message bénévole "+str(index_ben+1),"Le message n'a pas pu être envoyé ! Pensez à recharger votre compte Sinch : https://dashboard.sinch.com/dashboard")
                print('Failed to communicate with XMS: %s' % str(ex))

def changed():
    # A chaque fois qu'un bénéficiaires est coché (affecté au bénévole x), cette fonction s'active.
    # Mets à jours la liste des bénéficiaires restant pour les autres bénévoles (en supprimant celle coché).
    # Cette fonction fait donc appel à la fonction findDiff()
    
    current_dic = app.getAllProperties()
    global last_message
    global last_message_new
    global message
    global beneficiaires_par_benevole
    global previous_dic
    global benevoles
    global benevole
    findDiff(current_dic,previous_dic)
    previous_dic = current_dic
    index_ben = int(benevole[-2:])-1
    toppings = app.getAllProperties()['Bénéficiaires dédiés au bénévole '+str(index_ben+1)]
    benevole_name = app.getOptionBox("Bénévole "+str(index_ben+1))

    for item in toppings:
        boolean = False

        if toppings[item] == True:

            try:

                n = app.getTableRowCount("g"+str(index_ben+1))

                for i in range(0,n):
                    x = app.getTableRow("g"+str(index_ben+1),i)
                    if Convert(item) == x:
                        boolean = True # deja dans table

                if boolean is False:
                    app.addTableRow("g"+str(index_ben+1),Convert(item))
                    beneficiaire=Convert(item)
                    beneficiaires_par_benevole[index_ben].append(beneficiaire)
                    ##########
                    message[index_ben] = message[index_ben] + "- " +beneficiaires_par_benevole[index_ben][-1][0] + "\nsitué à : " +beneficiaires_par_benevole[index_ben][-1][1]+"\ndont le numéro de téléphone est : "+ beneficiaires_par_benevole[index_ben][-1][2]+",\n"
                    ##########
                    app.setLabel("l"+str(index_ben+1), message[index_ben])  
                    for k in range(0, len_ben):
                        if k != index_ben:
                            app.deleteProperty("Bénéficiaires dédiés au bénévole "+str(k+1), item)

            except:
                pass
        else:
            n = app.getTableRowCount("g"+str(index_ben+1))
            beneficiaire=Convert(item)
            message[index_ben]=message[index_ben].replace("\n- "+beneficiaire[0] + "\nsitué à : " +beneficiaire[1]+"\ndont le numéro de téléphone est : "+ beneficiaire[2]+",","")
            for beneficiares_ in beneficiaires_par_benevole:
                if beneficiaire in beneficiares_:
                    beneficiaires_par_benevole[index_ben].remove(beneficiaire)

            app.setLabel("l"+str(index_ben+1), message[index_ben])
            for i in range(0,n):
                try:
                    x = app.getTableRow("g"+str(index_ben+1),i)
                    if Convert(item) == x:
                        app.deleteTableRow("g"+str(index_ben+1), i)
                except:
                    pass
                    
            for k in range(0,len_ben):
                try:
                    app.setProperty("Bénéficiaires dédiés au bénévole "+str(k+1), item, callFunction=False)
                except:
                    pass
        if last_message[index_ben] != "":
            message[index_ben] = message[index_ben].replace(last_message[index_ben],"")
        message[index_ben] = message[index_ben] + last_message_new[index_ben]
        last_message[index_ben] = last_message_new[index_ben]

def changedBenevole():
    # Cette fonction se déclanche lorsqu'un désire envoyer un message à un nouveau bénévole.
    # Le nom du nouveau bénévole à qui on souhaite envoyer le message est changé !

    global benevoles
    global message
    global beneficiaires_par_benevole
    global last_message
    global last_message_new
    # message = ["Message à envoyer à "]*len(benevoles)
    for index_ben in range(0,len(benevoles)):
        benevole_name = app.getOptionBox("Bénévole "+str(index_ben+1))
        message[index_ben]=message[index_ben].replace( "Message à envoyer à "+benevoles[index_ben],"Message à envoyer à "+benevole_name)
        message[index_ben]=message[index_ben].replace( "Bonjour "+benevoles[index_ben], "Bonjour "+benevole_name)
        benevoles[index_ben]=benevole_name
        print("changed/n")
        print(last_message_new[index_ben])
        if last_message[index_ben] != "":
            message[index_ben]=message[index_ben].replace(last_message[index_ben],"")
        app.setLabel("l"+str(index_ben+1),message[index_ben]+last_message_new[index_ben])
        last_message[index_ben]=last_message_new[index_ben]

def changedPlace():
    # Fonction qui détecte les changements dans les références de la personne de contacte (+ date rdz et lieu de rdz).
    # Mets à jour le 'last_message' contenant la : 
    # - Le numéro de la personne de contacte
    # - Le lieu de rendez-vous
    # - La date de rendez-vous

    global numero_contact
    global last_message_new
    dictionary = app.getAllEntries()
    for item in dictionary:
        if "Numéro de contact" in item:
            numero_contact[int(item[-2:])-1] = dictionary[item]
        elif "Lieu rendez-vous" in item:
            lieu_rdz[int(item[-2:])-1]= dictionary[item]
        else:
            date_rdz[int(item[-2:])-1]= dictionary[item]
            last_message_new[int(item[-2:])-1] = "Pour toutes questions, tu peux m'appeler au " +numero_contact[int(item[-2:])-1]+".\n"+ "La date de rendez-vous est : "+date_rdz[int(item[-2:])-1]+ ". Le lieu de rendez-vous est : "+lieu_rdz[int(item[-2:])-1]+ ".\nPreviens-nous si tu as un retard."+ "\nA très bientôt !"
            print(numero_contact)
            print(last_message_new[0])
#####################################################################################
app = gui("Main tendue")
app.showSplash('Main tendue', fill='red', stripe='black', fg='white', font=100)
# app.setSize("Fullscreen")

app.setFont(10)
try:
    wb = load_workbook(filename = 'Liste de bénéficiaires.xlsx')
except:
    app.infoBox("Erreur", "Le fichier excel Liste de bénéficiaires n'a pas pu être chargé. Vérifiez que le fichier est présent dans le même dossier que l'application. Vérifiez aussi que le nom du fichier est bien : Liste de bénéficiaires", parent=None)
    app.stop()
ws = wb['Bénévoles']
global now_
global now
now = datetime.now()
now = str(now)
now_ = now.replace("-",".")
now_ = now.replace(":",';')
global phone_numbers_benev_dic
phone_numbers_benev_dic={}
phone_numbers_benev_list = []
global benevoles
benevoles = []
for r in range(2, ws.max_row + 1):
    if ws.cell(row = r, column = 1).value is not None:
        benevoles.append(ws.cell(row = r, column = 1).value)
        phone_numbers_benev_list.append(ws.cell(row = r, column = 2).value)
for (index_ben,benevole) in zip(range(0,len(benevoles)),benevoles):
    phone_numbers_benev_dic.update({benevole:phone_numbers_benev_list[index_ben]})
for i in range(0,len(benevoles)):
    print(benevoles[i]+" "+str(phone_numbers_benev_list[i])+"\n")

global numero_contact
numero_contact = ["None"]*len(benevoles)
global date_rdz
date_rdz = [""]*len(benevoles)
global lieu_rdz
lieu_rdz = [""]*len(benevoles)
ws = wb['Liste des bénéficiaires (Colis)']
beneficiaires = []
phone_numbers_benef = []
adresse_benef = []
nombre_personnes = []
for r in range(2, ws.max_row + 1):
    beneficiaires.append(ws.cell(row = r, column = 1).value)
    phone_numbers_benef.append(ws.cell(row = r, column = 3).value)
    adresse_benef.append(ws.cell(row = r, column = 4).value)
    nombre_personnes.append(ws.cell(row = r, column = 2).value)

beneficiaires = list(filter(None,beneficiaires))
phone_numbers_benef = list(filter(None,phone_numbers_benef))
adresse_benef = list(filter(None,adresse_benef))
nombre_personnes = list(filter(None,nombre_personnes))

global last_message 
global last_message_new
last_message = [""]*len(benevoles)
last_message_new = [""]*len(benevoles)
toppings={"Fred-45-Male":False, "Tina-37-Female":False, "Jean-37-Female":False, "Jeel-37-Female":False}
beneficiaires_dic = {}
for i in range(0,len(beneficiaires)):
    beneficiaires_dic.update({str(beneficiaires[i]) + "-" + str(adresse_benef[i]) + "-" + str(phone_numbers_benef[i]) + "-" + str(nombre_personnes[i]):False})
global len_ben
global message
len_ben = len(benevoles)
app.startTabbedFrame('TabbedFrame')
global beneficiaires_par_benevole
introduction = "Bonjour "
message = ["Message à envoyer à "]*len_ben
beneficiaires_par_benevole = [[]]*len_ben
for index_ben in range(0,len(benevoles)):
    app.startTab("Bénévole "+str(index_ben+1))
    
    message[index_ben]=message[index_ben]+str(benevoles[index_ben])+ ":\n\n"
    message[index_ben] = message[index_ben] + "Bonjour " + benevoles[index_ben] +",\n"+ "Comme convenu, les bénéficiares qui te seront assignés sont :\n"
#########
    app.startFrame("LEFT"+str(index_ben), row=0, column=0)
    
    app.startScrollPane("l"+str(index_ben+1),disabled="horizontal")
    app.addLabel("Réferences bénévole "+str(index_ben+1),"Réferences bénévole "+str(index_ben+1))
    app.addLabelEntry("Numéro de contact " + str(index_ben+1))
    app.setEntry("Numéro de contact "+ str(index_ben+1), "+320 000 000",callFunction=True)
    app.setEntryChangeFunction("Numéro de contact "+ str(index_ben+1), changedPlace)
    app.addLabelEntry("Lieu rendez-vous " + str(index_ben+1))
    app.setEntry("Lieu rendez-vous "+ str(index_ben+1), "Lieu rendez-vous",callFunction=True)
    app.setEntryChangeFunction("Lieu rendez-vous "+ str(index_ben+1), changedPlace)
    app.addLabelEntry("Date rendez-vous " + str(index_ben+1))
    app.setEntry("Date rendez-vous "+ str(index_ben+1), "Dimanche xx/xx/xxxx",callFunction=True)
    app.setEntryChangeFunction("Date rendez-vous "+ str(index_ben+1), changedPlace)
    app.addLabel("l"+str(index_ben+1),message[index_ben])
    app.addButton("Envoyer un message au bénévole "+str(index_ben+1), press)
    app.addEmptyMessage("Message bénévole "+str(index_ben+1))
    app.setMessageWidth("Message bénévole "+str(index_ben+1), 300)
    app.stopScrollPane()
    app.stopFrame()

#########
    app.startFrame("RIGHT"+str(index_ben), row=0, column=1)
    app.addLabelOptionBox("Bénévole "+str(index_ben+1), benevoles)
    app.setOptionBox("Bénévole "+str(index_ben+1), index_ben, value=True, callFunction=True, override=False)
    app.setOptionBoxChangeFunction("Bénévole "+str(index_ben+1), changedBenevole)
    app.startScrollPane("Bénéficiaires dédiés au bénévole "+str(index_ben+1),disabled="horizontal")
    app.addProperties("Bénéficiaires dédiés au bénévole "+str(index_ben+1), beneficiaires_dic)
    app.setPropertiesChangeFunction("Bénéficiaires dédiés au bénévole "+str(index_ben+1), changed)
    app.stopScrollPane()
    app.addTable("g"+str(index_ben+1),
        [["Nom", "Adresse", "Numéro", "Nombre de personnes"]], showMenu=True)
#########
    app.stopFrame()
    app.stopTab()
app.stopTabbedFrame()
global previous_dic
previous_dic = app.getAllProperties()
app.addWebLink("Cliquez ici pour consulter votre solde Sinch", "https://dashboard.sinch.com/dashboard")
app.go()